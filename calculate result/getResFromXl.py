import openpyxl

# Open the first XLSX file
wb1 = openpyxl.load_workbook('labels1.xlsx')
ws1 = wb1.active

# Open the second XLSX file
wb2 = openpyxl.load_workbook('res1.xlsx')
ws2 = wb2.active

sum_of_non_anomaly = 0
non_anomaly_predicted = 0
non_anomaly_predicted_wrong = 0

sum_of_anomaly = 0
anomaly_predicted = 0
anomaly_predicted_wrong = 0

for i in range(1, ws1.max_row):
    tr_res_tmp = str((ws1.cell(row=i, column=2).value))
    our_res_tmp = str((ws2.cell(row=i, column=2).value))

    if tr_res_tmp == "0":
        sum_of_non_anomaly += 1
        if our_res_tmp == "0":
            non_anomaly_predicted += 1
        else:
            non_anomaly_predicted_wrong += 1

    if tr_res_tmp == "1":
        sum_of_anomaly += 1
        if our_res_tmp == "1":
            anomaly_predicted += 1
        else:
            anomaly_predicted_wrong += 1

print("result:")

print("sum_of_non_anomaly = ", sum_of_non_anomaly)
print("non_anomaly_predicted = ", non_anomaly_predicted)
print("non_anomaly_predicted_wrong = ", non_anomaly_predicted_wrong)
print()
print("sum_of_anomaly = ", sum_of_anomaly)
print("anomaly_predicted = ", anomaly_predicted)
print("anomaly_predicted_wrong = ", anomaly_predicted_wrong)
print()
print("accuracy = (non_anomaly_predicted + anomaly_predicted) / (sum_of_non_anomaly + sum_of_anomaly)")
print("accuracy = ", (non_anomaly_predicted + anomaly_predicted) / (sum_of_non_anomaly + sum_of_anomaly)*100, "%")
print()
print("recall = anomaly_predicted / sum_of_anomaly")
print("recall = ", (anomaly_predicted / sum_of_anomaly)*100, "%")
print()
print("Confusion matrix and accuracy")
print("                   | non_anomaly_predicted | anomaly_predicted  ")
print(f"actual non anomaly |         {non_anomaly_predicted}        |         {non_anomaly_predicted_wrong}       ")
print(f"actual anomaly     |           {anomaly_predicted_wrong}           |         {anomaly_predicted}       ")
